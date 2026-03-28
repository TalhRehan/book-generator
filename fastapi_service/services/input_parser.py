"""Service-layer business logic for the book generation workflow."""

import openpyxl
from pathlib import Path
from fastapi_service.api.schemas import BookInputRow


EXPECTED_COLUMNS = [
    "title",
    "notes_on_outline_before",
    "notes_on_outline_after",
    "status_outline_notes",
    "final_review_notes_status",
]


def parse_excel(file_path: str) -> list[BookInputRow]:
    """Parse excel."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {file_path}")

    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    missing = [col for col in EXPECTED_COLUMNS if col not in headers]
    if missing:
        raise ValueError(f"Missing columns in Excel: {missing}")

    col_index = {name: headers.index(name) for name in EXPECTED_COLUMNS}
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        title = row[col_index["title"]]
        if not title:
            continue

        rows.append(BookInputRow(
            title=str(title).strip(),
            notes_on_outline_before=_clean(row[col_index["notes_on_outline_before"]]),
            notes_on_outline_after=_clean(row[col_index["notes_on_outline_after"]]),
            status_outline_notes=_clean(row[col_index["status_outline_notes"]]) or "no",
            final_review_notes_status=_clean(row[col_index["final_review_notes_status"]]) or "no",
        ))

    return rows


def _clean(value) -> str | None:
    """Clean."""
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned if cleaned else None